import xlsxwriter as xlsxwriter
from nltk import wordpunct_tokenize
from nltk.corpus import stopwords
from openpyxl import load_workbook
import os
import re
from openpyxl import load_workbook
from collections import Counter
import enchant
import unicodedata
import langdetect
forumdataforreading = "datafiles\datatoclean.xlsx"



def excelinput(filetoeopn, filecheckksheets, columnNo):
    columnlist = []
    # loading in workbook
    wb = load_workbook(filetoeopn)

    print("Sheets names:")
    # obtaining sheets names
    print(wb.get_sheet_names())
    sheet = wb[wb.get_sheet_names()[filecheckksheets]]

    #
    data = sheet.values

    for r in data:
        # print(r[0])
        if r[columnNo] is not None:
            # print(r[columnNo])
            columnlist.append(str(r[columnNo]))

    return columnlist

counttttt = 0
f = open('cleanedData.txt', 'w', encoding="utf-8")
# RAW DATA HERE
for i in excelinput(forumdataforreading, 0, 0):
    # print(i)
    print("Currently scanning Line: " + str(counttttt))
    counttttt = counttttt + 1
    i = str(i).replace('\n', " ")
    i = str(i).replace('\\n', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")

    i = str(i).replace('…', '...')

    # i = str(i).replace('/', " / ")
    # i = str(i).replace('"', " ")
    # i = str(i).replace('(', " ")
    # i = str(i).replace(')', " ")
    # i = str(i).replace('“', " ")
    # i = str(i).replace('. ', " ")
    # i = str(i).replace(';', " ")
    # i = str(i).replace('[', " ")
    # i = str(i).replace(']', " ")
    # i = str(i).lower()
    # do blank check to see if have space

    print(i)
    f.writelines(str(i)+"\n")

f.close()