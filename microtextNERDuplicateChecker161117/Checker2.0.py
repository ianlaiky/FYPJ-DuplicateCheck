import xlsxwriter
from openpyxl import load_workbook
from collections import Counter

filesToReadList = []

fileconfigread = open("checkerConfig.txt", 'r', encoding="utf-8")

for fsfsd in fileconfigread:
    if str(fsfsd.strip()) != "":
        if str(fsfsd.strip()) != " ":

            if str(fsfsd[:1]) != "#":
                print(str(fsfsd).strip())

                filesToReadList.append(str(fsfsd).strip())

print(filesToReadList)


# tempcurrSheet=""

def excelinput(filetoeopn, filecheckksheets, columnNo):
    columnlist = []
    # loading in workbook
    wb = load_workbook(filetoeopn)

    print("Sheets names:")
    # obtaining sheets names
    print(wb.get_sheet_names())
    print("Sheets Loaded: " + wb.get_sheet_names()[int(filecheckksheets)])
    sheet = wb[wb.get_sheet_names()[int(filecheckksheets)]]

    # tempcurrSheet=str(sheet)

    #
    data = sheet.values

    # print(next(data)[0:])

    for r in data:
        # print(r)

        if str(r) != "":
            if str(r) != " ":
                if str(str(r[int(columnNo)])) != "None":
                    print("Looking through.... Please Wait: " + str(r[int(columnNo)]))
                # print(r[0])

                columnlist.append(str(r[int(columnNo)]))
    print("Saving...Please Wait. This may take a while")
    return columnlist


class Data():
    def __init__(self, col0, col1, col2, col3):
        self.savcol0 = col0
        self.savcol1 = col1
        self.savcol2 = col2
        self.savcol3 = col3

    def getcol0(self):
        return self.savcol0

    def getcol1(self):
        return self.savcol1

    def getcol2(self):
        return self.savcol2

    def getcol3(self):
        return self.savcol3

    def setcol0(self, scol0):
        self.savcol0 = scol0

    def setcol1(self, scol1):
        self.savcol1 = scol1

    def setcol2(self, scol2):
        self.savcol2 = scol2

    def setcol3(self, scol3):
        self.savcol3 = scol3


listarrayOfFilesData = []
listOfAllAbbr = []

dictOfObj = {}

for readfiles in filesToReadList:
    arrayOfFilesData = []
    wb = load_workbook(str(readfiles).strip().split(",")[0])
    print(wb.get_sheet_names()[int(str(readfiles).strip().split(",")[1])])
    print(str(readfiles).strip().split(",")[0])
    print(str(readfiles).strip().split(",")[1])
    print(str(readfiles).strip().split(",")[2])
    print(str(readfiles).strip().split(",")[3])
    print(str(readfiles).strip().split(",")[4])
    print("Loading files.. Please wait.")
    colAbbr = excelinput(str(readfiles).strip().split(",")[0], str(readfiles).strip().split(",")[1],
                         str(readfiles).strip().split(",")[2])
    colFullForm = excelinput(str(readfiles).strip().split(",")[0], str(readfiles).strip().split(",")[1],
                             str(readfiles).strip().split(",")[3])
    colPolar = excelinput(str(readfiles).strip().split(",")[0], str(readfiles).strip().split(",")[1],
                          str(readfiles).strip().split(",")[4])

    sheetnames = wb.get_sheet_names()[int(str(readfiles).strip().split(",")[1])]

    FinalcolAbbr = []
    FinalcolFullForm = []
    FinalcolcolPolar = []
    FinalSheetName = []

    for index, checkcol in enumerate(colAbbr):
        if str(checkcol) != "":
            if str(checkcol) != " ":
                if str(checkcol) != "None":
                    listOfAllAbbr.append(str(colAbbr[index]).lower())
                    FinalcolAbbr.append(str(colAbbr[index]))
                    FinalcolFullForm.append(str(colFullForm[index]))
                    FinalcolcolPolar.append(str(colPolar[index]))
                    FinalSheetName.append(str(sheetnames))

    arrayOfFilesData.append(FinalcolAbbr)
    arrayOfFilesData.append(FinalcolFullForm)
    arrayOfFilesData.append(FinalcolcolPolar)
    arrayOfFilesData.append(FinalSheetName)
    listarrayOfFilesData.append(arrayOfFilesData)

freqCount = Counter(listOfAllAbbr)

microtextEnglishFreqMore2 = []

for counts in freqCount:
    if int(freqCount[counts]) > 1:
        microtextEnglishFreqMore2.append(str(counts).lower())
print(listarrayOfFilesData)

for fdsdsfs in listarrayOfFilesData:
    print(fdsdsfs)

# print(listarrayOfFilesData[0])
# print(listarrayOfFilesData[0][0])
# print(listarrayOfFilesData[0][0][0])
# input()


for index2, dasdf in enumerate(listarrayOfFilesData):
    print("Checking:")
    print(dasdf)

    for index13, readin in enumerate(dasdf[0]):

        # print("prinnting")
        # print(readin)
        # input()
        if str(readin).lower() in microtextEnglishFreqMore2:
            # if str(readin).lower() in microtextEnglishFreqMore2:
                # if str(readin).lower()=="ai":
                #     countdsdsa=int(countdsdsa)+1
            if str(readin).lower() in dictOfObj:
                # print("wrer")

                getcurrObj = dictOfObj[str(readin).lower()]
                currCol0 = str(getcurrObj.getcol0())
                currCol1 = str(getcurrObj.getcol1())
                currCol2 = str(getcurrObj.getcol2())
                currCol3 = str(getcurrObj.getcol3())

                # print(currCol0)
                # print(currCol1)
                # print(currCol2)
                # print(currCol3)

                newCol1 = str(currCol1) + ", " + str(listarrayOfFilesData[index2][1][index13])
                newCol2 = str(currCol2) + ", " + str(listarrayOfFilesData[index2][2][index13])
                newCol3 = str(currCol3) + ", " + str(listarrayOfFilesData[index2][3][index13])

                dictOfObj[str(readin).lower()] = Data(str(currCol0), str(newCol1), str(newCol2), str(newCol3))

            else:
                # print(listarrayOfFilesData)
                # print(listarrayOfFilesData[index2])
                # print(listarrayOfFilesData[index2][0])
                # print(readin)

                # print(listarrayOfFilesData[index2][index13][0])
                # print(listarrayOfFilesData[index2][index13][0][index1])
                # print(str(listarrayOfFilesData[index2][1][index13]))
                # print(str(listarrayOfFilesData[index2][2][index13]))

                dictOfObj[str(readin).lower()] = Data(str(listarrayOfFilesData[index2][0][index13]),
                                                      str(listarrayOfFilesData[index2][1][index13]),
                                                      str(listarrayOfFilesData[index2][2][index13]),
                                                      str(listarrayOfFilesData[index2][3][index13]))

                # print("fdsdsfdsfdsfdsfds")

row1 = 1

workbook1 = xlsxwriter.Workbook('DuplicatedDataFile.xlsx')
worksheet1 = workbook1.add_worksheet()

worksheet1.write(0, 0, "Overlap")
worksheet1.write(0, 1, "Full Form (Microtext/NER)")
worksheet1.write(0, 2, "Detail (Polarity/NER Category)")
worksheet1.write(0, 3, "Source (Tab)")

# print(dictOfObj["12"].getcol0)

for uiui in dictOfObj:
    # print(dictOfObj[uiui].getcol0())
    # print(dictOfObj[uiui].getcol1())
    # print(dictOfObj[uiui].getcol2())
    # print(dictOfObj[uiui].getcol3())
    worksheet1.write(row1, 0, dictOfObj[uiui].getcol0())
    worksheet1.write(row1, 1, dictOfObj[uiui].getcol1())
    worksheet1.write(row1, 2, dictOfObj[uiui].getcol2())
    worksheet1.write(row1, 3, dictOfObj[uiui].getcol3())
    row1 = row1 + 1

workbook1.close()
print("Overlapped Found:" + str(len(microtextEnglishFreqMore2)))

print("AIII0" + str(freqCount["ai"]))
