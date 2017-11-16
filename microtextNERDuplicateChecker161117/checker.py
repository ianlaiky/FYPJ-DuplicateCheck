import xlsxwriter
from openpyxl import load_workbook
from collections import Counter
import fnmatch


def excelinput(filetoeopn, filecheckksheets, columnNo):
    columnlist = []
    # loading in workbook
    wb = load_workbook(filetoeopn)

    print("Sheets names:")
    # obtaining sheets names
    print(wb.get_sheet_names())
    print("Sheets Loaded: " + wb.get_sheet_names()[filecheckksheets])
    sheet = wb[wb.get_sheet_names()[filecheckksheets]]

    #
    data = sheet.values

    # print(next(data)[0:])

    for r in data:

        if str(r) != "":
            if str(r) != " ":
                # print(r[0])
                columnlist.append(r[columnNo])

    return columnlist


# col 1 to check
microTextEnglishFull = []
microTextSinglishFull = []
nerLocationFull = []
nerOrganisationFull = []
nerPeopleFull = []
nerMiscellaneousFull = []

microTextEnglishCol1 = excelinput("dataFiles\Microtext.xlsx", 2, 0)
microTextEnglishCol2 = excelinput("dataFiles\Microtext.xlsx", 2, 1)
microTextEnglishCol3 = excelinput("dataFiles\Microtext.xlsx", 2, 2)

del microTextEnglishCol1[0]
del microTextEnglishCol2[0]
del microTextEnglishCol3[0]

microTextEnglishFull.append(microTextEnglishCol1)
microTextEnglishFull.append(microTextEnglishCol2)
microTextEnglishFull.append(microTextEnglishCol3)
# print(microTextEnglishFull)

microTextSinglishCol1 = excelinput("dataFiles\Microtext.xlsx", 0, 0)
microTextSinglishCol2 = excelinput("dataFiles\Microtext.xlsx", 0, 1)
microTextSinglishCol3 = excelinput("dataFiles\Microtext.xlsx", 0, 2)

del microTextSinglishCol1[0]
del microTextSinglishCol2[0]
del microTextSinglishCol3[0]

microTextSinglishFull.append(microTextSinglishCol1)
microTextSinglishFull.append(microTextSinglishCol2)
microTextSinglishFull.append(microTextSinglishCol3)

nerLocationCol1 = excelinput("dataFiles\\NER.xlsx", 5, 3)
nerLocationCol2 = excelinput("dataFiles\\NER.xlsx", 5, 0)
nerLocationCol3 = excelinput("dataFiles\\NER.xlsx", 5, 1)

del nerLocationCol1[0]
del nerLocationCol2[0]
del nerLocationCol3[0]

newnerLocationCol1 = []
newnerLocationCol2 = []
newnerLocationCol3 = []

# print(nerLocationCol1)


for index, ner1 in enumerate(nerLocationCol1):

    if str(ner1) != "None":
        newnerLocationCol1.append(nerLocationCol1[index])
        newnerLocationCol2.append(nerLocationCol2[index])
        newnerLocationCol3.append(nerLocationCol3[index])

nerLocationFull.append(newnerLocationCol1)
nerLocationFull.append(newnerLocationCol2)
nerLocationFull.append(newnerLocationCol3)

nerOrganisationCol1 = excelinput("dataFiles\\NER.xlsx", 6, 3)
nerOrganisationCol2 = excelinput("dataFiles\\NER.xlsx", 6, 0)
nerOrganisationCol3 = excelinput("dataFiles\\NER.xlsx", 6, 1)

del nerOrganisationCol1[0]
del nerOrganisationCol2[0]
del nerOrganisationCol3[0]

newOrganisationCol1 = []
newOrganisationCol2 = []
newOrganisationCol3 = []

for indexNer2, ner2 in enumerate(nerOrganisationCol1):
    if str(ner2) != "None":
        newOrganisationCol1.append(nerOrganisationCol1[indexNer2])
        newOrganisationCol2.append(nerOrganisationCol2[indexNer2])
        newOrganisationCol3.append(nerOrganisationCol3[indexNer2])

nerOrganisationFull.append(newOrganisationCol1)
nerOrganisationFull.append(newOrganisationCol2)
nerOrganisationFull.append(newOrganisationCol3)

nerPeopleCol1 = excelinput("dataFiles\\NER.xlsx", 7, 3)
nerPeopleCol2 = excelinput("dataFiles\\NER.xlsx", 7, 0)
nerPeopleCol3 = excelinput("dataFiles\\NER.xlsx", 7, 1)

del nerPeopleCol1[0]
del nerPeopleCol2[0]
del nerPeopleCol3[0]

newPeopleCol1 = []
newPeopleCol2 = []
newPeopleCol3 = []

for indexner3, ner3 in enumerate(nerPeopleCol1):
    if str(ner3) != "None":
        newPeopleCol1.append(nerPeopleCol1[indexner3])
        newPeopleCol2.append(nerPeopleCol2[indexner3])
        newPeopleCol3.append(nerPeopleCol3[indexner3])

nerPeopleFull.append(newPeopleCol1)
nerPeopleFull.append(newPeopleCol2)
nerPeopleFull.append(newPeopleCol3)

# print(nerPeopleFull[0][5])
# print(nerPeopleFull[1][5])
# print(nerPeopleFull[2][5])

nerMiscellaneousCol1 = excelinput("dataFiles\\NER.xlsx", 8, 3)
nerMiscellaneousCol2 = excelinput("dataFiles\\NER.xlsx", 8, 0)
nerMiscellaneousCol3 = excelinput("dataFiles\\NER.xlsx", 8, 1)

del nerMiscellaneousCol1[0]
del nerMiscellaneousCol2[0]
del nerMiscellaneousCol3[0]

newNerMiscellaneousCol1 = []
newNerMiscellaneousCol2 = []
newNerMiscellaneousCol3 = []

for indexner4, ner4 in enumerate(nerMiscellaneousCol1):
    if str(ner4) != "None":
        newNerMiscellaneousCol1.append(nerMiscellaneousCol1[indexner4])
        newNerMiscellaneousCol2.append(nerMiscellaneousCol2[indexner4])
        newNerMiscellaneousCol3.append(nerMiscellaneousCol3[indexner4])

nerMiscellaneousFull.append(newNerMiscellaneousCol1)
nerMiscellaneousFull.append(newNerMiscellaneousCol2)
nerMiscellaneousFull.append(newNerMiscellaneousCol3)

# print(nerMiscellaneousFull[0][5])
# print(nerMiscellaneousFull[1][5])
# print(nerMiscellaneousFull[2][5])


microtextEnglishFreqMore2 = []
alllist = []
alllist = microTextEnglishFull[0]
alllist = alllist + microTextSinglishFull[0]
alllist = alllist + nerLocationFull[0]
alllist = alllist + nerOrganisationFull[0]
alllist = alllist + nerPeopleFull[0]
alllist = alllist + nerMiscellaneousFull[0]

listOfList = []
listOfList.append(microTextEnglishFull)
listOfList.append(microTextSinglishFull)
listOfList.append(nerLocationFull)
listOfList.append(nerOrganisationFull)
listOfList.append(nerPeopleFull)
listOfList.append(nerMiscellaneousFull)

freqCount = Counter(alllist)

print(alllist)
print(Counter(alllist))

for counts in freqCount:
    if int(freqCount[counts]) > 1:
        microtextEnglishFreqMore2.append(str(counts).lower())

print(microtextEnglishFreqMore2)


class Data():
    def __init__(self, col0, col1, col2, col3, col4, col5, col6):
        self.savcol0 = col0
        self.savcol1 = col1
        self.savcol2 = col2
        self.savcol3 = col3
        self.savcol4 = col4
        self.savcol5 = col5
        self.savcol6 = col6

    def getcol0(self):
        return self.savcol0

    def getcol1(self):
        return self.savcol1

    def getcol2(self):
        return self.savcol2

    def getcol3(self):
        return self.savcol3

    def getcol4(self):
        return self.savcol4

    def getcol5(self):
        return self.savcol5

    def getcol6(self):
        return self.savcol6

    def setcol0(self, scol0):
        self.savcol0 = scol0

    def setcol1(self, scol1):
        self.savcol1 = scol1

    def setcol2(self, scol2):
        self.savcol2 = scol2

    def setcol3(self, scol3):
        self.savcol3 = scol3

    def setcol4(self, scol4):
        self.savcol4 = scol4

    def setcol5(self, scol5):
        self.savcol5 = scol5

    def setcol6(self, scol6):
        self.savcol6 = scol6


savingasAdict = {}

for indexCurr, scroll1 in enumerate(microTextEnglishFull[0]):

    if str(scroll1).lower() in microtextEnglishFreqMore2:
        if str(scroll1).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll1)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col1OfObj) != "":
                newCol1 = str(col1OfObj) + ", " + str(microTextEnglishFull[1][int(indexCurr)])
            else:
                newCol1 = str(microTextEnglishFull[1][int(indexCurr)])
            if str(col2OfObj) != "":
                newCol2 = str(col2OfObj) + ", " + str(microTextEnglishFull[2][int(indexCurr)])
            else:
                newCol2 = str(microTextEnglishFull[2][int(indexCurr)])

            # setting new data
            savingasAdict[str(scroll1).lower()] = Data(col0OfObj, newCol1, newCol2, col3OfObj, col4OfObj,
                                               col5OfObj, col6OfObj)


        else:
            savingasAdict[str(scroll1).lower()] = Data(str(microTextEnglishFull[0][int(indexCurr)]),
                                               str(microTextEnglishFull[1][int(indexCurr)]),
                                               str(microTextEnglishFull[2][int(indexCurr)]), "", "", "", "")

        print(str(microTextEnglishFull[0][int(indexCurr)]))

for indexCurr1, scroll2 in enumerate(microTextSinglishFull[0]):
    if str(scroll2).lower() in microtextEnglishFreqMore2:
        if str(scroll2).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll2)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col3OfObj) != "":
                newcol3 = str(col3OfObj) + ", " + str(microTextSinglishFull[1][int(indexCurr1)])
            else:
                newcol3 = str(microTextSinglishFull[1][int(indexCurr1)])

            if str(col4OfObj) != "":
                newcol4 = str(col4OfObj) + ", " + str(microTextSinglishFull[2][int(indexCurr1)])
            else:
                newcol4 = str(microTextSinglishFull[2][int(indexCurr1)])

            # setting new data
            savingasAdict[str(scroll2).lower()] = Data(col0OfObj, col1OfObj, col2OfObj, newcol3, newcol4, col5OfObj, col6OfObj)
        else:
            savingasAdict[str(scroll2).lower()] = Data(microTextSinglishFull[0][int(indexCurr1)], "", "",
                                               microTextSinglishFull[1][int(indexCurr1)],
                                               microTextSinglishFull[2][int(indexCurr1)], "", "")

for indexCurr2, scroll3 in enumerate(nerLocationFull[0]):
    if str(scroll3).lower() in microtextEnglishFreqMore2:
        if str(scroll3).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll3)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col5OfObj) != "":
                newcol5 = str(col5OfObj) + ", " + str(nerLocationFull[1][int(indexCurr2)])
            else:
                newcol5 = str(nerLocationFull[1][int(indexCurr2)])
            if str(col6OfObj) != "":
                newcol6 = str(col6OfObj) + ", " + str(nerLocationFull[2][int(indexCurr2)])
            else:
                newcol6 = str(nerLocationFull[2][int(indexCurr2)])

            savingasAdict[str(scroll3).lower()] = Data(col0OfObj, col1OfObj, col2OfObj, col3OfObj, col4OfObj, newcol5, newcol6)
        else:
            savingasAdict[str(scroll3).lower()] = Data(nerLocationFull[0][int(indexCurr2)], "", "", "", "",
                                               nerLocationFull[1][int(indexCurr2)], nerLocationFull[2][int(indexCurr2)])

for indexCurr2, scroll3 in enumerate(nerOrganisationFull[0]):
    if str(scroll3).lower() in microtextEnglishFreqMore2:
        if str(scroll3).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll3)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col5OfObj) != "":
                newcol5 = str(col5OfObj) + ", " + str(nerOrganisationFull[1][int(indexCurr2)])
            else:
                newcol5 = str(nerOrganisationFull[1][int(indexCurr2)])

            if str(col6OfObj) != "":
                newcol6 = str(col6OfObj) + ", " + str(nerOrganisationFull[2][int(indexCurr2)])
            else:
                newcol6 = str(nerOrganisationFull[2][int(indexCurr2)])

            savingasAdict[str(scroll3).lower()] = Data(col0OfObj, col1OfObj, col2OfObj, col3OfObj, col4OfObj, newcol5, newcol6)
        else:
            savingasAdict[str(scroll3).lower()] = Data(nerOrganisationFull[0][int(indexCurr2)], "", "", "", "",
                                               nerOrganisationFull[1][int(indexCurr2)],
                                               nerOrganisationFull[2][int(indexCurr2)])

for indexCurr2, scroll3 in enumerate(nerPeopleFull[0]):
    if str(scroll3).lower() in microtextEnglishFreqMore2:
        if str(scroll3).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll3)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col5OfObj) != "":
                newcol5 = str(col5OfObj) + ", " + str(nerPeopleFull[1][int(indexCurr2)])
            else:
                newcol5 = str(nerPeopleFull[1][int(indexCurr2)])
            if str(col6OfObj) != "":
                newcol6 = str(col6OfObj) + ", " + str(nerPeopleFull[2][int(indexCurr2)])
            else:
                newcol6 = str(nerPeopleFull[2][int(indexCurr2)])

            savingasAdict[str(scroll3).lower()] = Data(col0OfObj, col1OfObj, col2OfObj, col3OfObj,
                                               col4OfObj,
                                               newcol5, newcol6)
        else:
            savingasAdict[str(scroll3).lower()] = Data(nerPeopleFull[0][int(indexCurr2)], "", "", "", "",
                                               nerPeopleFull[1][int(indexCurr2)],
                                               nerPeopleFull[2][int(indexCurr2)])

for indexCurr2, scroll3 in enumerate(nerMiscellaneousFull[0]):
    if str(scroll3).lower() in microtextEnglishFreqMore2:
        if str(scroll3).lower() in savingasAdict:
            objectt = savingasAdict[str(scroll3)]
            col0OfObj = objectt.getcol0()
            col1OfObj = objectt.getcol1()
            col2OfObj = objectt.getcol2()
            col3OfObj = objectt.getcol3()
            col4OfObj = objectt.getcol4()
            col5OfObj = objectt.getcol5()
            col6OfObj = objectt.getcol6()

            if str(col5OfObj) != "":
                newcol5 = str(col5OfObj) + ", " + str(nerMiscellaneousFull[1][int(indexCurr2)])
            else:
                newcol5 = str(nerMiscellaneousFull[1][int(indexCurr2)])
            if str(col6OfObj) != "":
                newcol6 = str(col6OfObj) + ", " + str(nerMiscellaneousFull[2][int(indexCurr2)])
            else:
                newcol6 = str(nerMiscellaneousFull[2][int(indexCurr2)])

            savingasAdict[str(scroll3).lower()] = Data(col0OfObj, col1OfObj, col2OfObj, col3OfObj, col4OfObj, newcol5, newcol6)
        else:
            savingasAdict[str(scroll3).lower()] = Data(nerMiscellaneousFull[0][int(indexCurr2)], "", "", "", "",
                                               nerMiscellaneousFull[1][int(indexCurr2)],
                                               nerMiscellaneousFull[2][int(indexCurr2)])

row1 = 1
col1 = 0

workbook1 = xlsxwriter.Workbook('DuplicatedData.xlsx')
worksheet1 = workbook1.add_worksheet()

print(len(savingasAdict))

worksheet1.write(0, 0, "Overlap")
worksheet1.write(0, 1, "MicroText_English")
worksheet1.write(0, 2, "Polarity_English")
worksheet1.write(0, 3, "Microtext_Singlish")
worksheet1.write(0, 4, "Polarity_Singlish")
worksheet1.write(0, 5, "NER")
worksheet1.write(0, 6, "Tag")

for ioio in savingasAdict:
    print(ioio)
    print(savingasAdict[ioio].getcol0())
    print(savingasAdict[ioio].getcol1())
    print(savingasAdict[ioio].getcol2())
    print(savingasAdict[ioio].getcol3())
    print(savingasAdict[ioio].getcol4())
    print(savingasAdict[ioio].getcol5())
    print(savingasAdict[ioio].getcol6())
    print("____________")

    worksheet1.write(row1, 0, savingasAdict[ioio].getcol0())
    worksheet1.write(row1, 1, savingasAdict[ioio].getcol1())
    worksheet1.write(row1, 2, savingasAdict[ioio].getcol2())
    worksheet1.write(row1, 3, savingasAdict[ioio].getcol3())
    worksheet1.write(row1, 4, savingasAdict[ioio].getcol4())
    worksheet1.write(row1, 5, savingasAdict[ioio].getcol5())
    worksheet1.write(row1, 6, savingasAdict[ioio].getcol6())

    row1 = row1 + 1

workbook1.close()
