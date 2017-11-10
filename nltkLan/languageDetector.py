import xlsxwriter as xlsxwriter
from nltk import wordpunct_tokenize
from nltk.corpus import stopwords
from openpyxl import load_workbook
import os
import re
from openpyxl import load_workbook
from collections import Counter
import enchant
import unicodedata


from langdetect import detect, DetectorFactory
from langdetect import detect_langs
DetectorFactory.seed = 0



def excelinput(filetoeopn, filecheckksheets, columnNo):
    columnlist = []
    # loading in workbook
    wb = load_workbook(filetoeopn)

    print("Sheets names:")
    # obtaining sheets names
    print(wb.get_sheet_names())
    sheet = wb[wb.get_sheet_names()[filecheckksheets]]

    #
    data = sheet.values

    for r in data:
        # print(r[0])
        if r[columnNo] is not None:
            # print(r[columnNo])
            columnlist.append(str(r[columnNo]))

    return columnlist
sentenceCount = 0

sgenCount=0

fsaveSinglish=open("SinglishSentences.txt",'w',encoding="utf-8")
ffailedDetect=open("FailedSentences.txt",'w',encoding="utf-8")

for i in excelinput("..\datafiles\Edmwcompiled311017.xlsx",0,0):

    i = str(i).replace('\n', " ")
    i = str(i).replace('\\n', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    i = str(i).replace('', " ")
    print(str(i).strip())
    # print(detect(i))

    print("Sentence No: "+str(sentenceCount))
    try:

        if str(detect(str(i).strip()))=="sgen":
            fsaveSinglish.writelines(str(i).strip()+"\n\n")
            sgenCount=sgenCount+1
    except:
        ffailedDetect.writelines(str(i).strip()+"\n\n")
        pass



    sentenceCount=sentenceCount+1
print("Singlish sentences: "+str(sgenCount)+"/"+str(sentenceCount))




print(detect('Have you eaten?'))
print(detect_langs('Dont be in like this way'))

fsaveSinglish.close()
ffailedDetect.close()