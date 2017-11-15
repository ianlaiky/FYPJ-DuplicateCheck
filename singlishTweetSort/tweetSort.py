import xlsxwriter as xlsxwriter
from nltk import wordpunct_tokenize
from nltk.corpus import stopwords
from openpyxl import load_workbook
import os
import re
from openpyxl import load_workbook
from collections import Counter
import enchant
import unicodedata

from langdetect import detect, DetectorFactory
from langdetect import detect_langs

DetectorFactory.seed = 0

sentenceCount = 0

sgenCount = 0

inputsen = open('post_unique.txt', 'r', encoding='latin-1')
failedsen = open('NonSinglishSentences.txt', 'w', encoding='utf-8')
singlishsen = open('SinglishSentences.txt', 'w', encoding='utf-8')

languages_ratios = {}
row1 = 0
col1 = 0

workbook1 = xlsxwriter.Workbook('SinglishSentences.xlsx')
worksheet1 = workbook1.add_worksheet()

sencoun = 0

for i in inputsen:
    if str(i) != "":
        if str(i) != " ":

            # print(str(i).strip())
            print("Sentences No: " + str(sencoun))

            try:

                if str(detect(str(i).strip())) == "sgen":
                    worksheet1.write(row1, col1, str(i))
                    singlishsen.writelines(str(i).strip() + "\n")
                    print(str(i))
                    row1 = row1 + 1
                else:
                    failedsen.writelines(str(i).strip() + "\n\n")
            except:
                failedsen.writelines(str(i).strip() + "\n\n")

    sencoun = sencoun + 1

workbook1.close()
singlishsen.close()
failedsen.close()
inputsen.close()
