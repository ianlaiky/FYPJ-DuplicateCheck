import re

import os
from openpyxl import load_workbook

import fnmatch

# import enchant
#
# d = enchant.Dict("en_US")
# de = enchant.Dict("en_GB")
#
# print(d.check("cool"))
# print(de.check("on-line"))
# print(de.check("afk"))
# input()

# if d.check("running") is True:
#     print("ts")

# import unicodedata
#
# def normalize_caseless(text):
#     return unicodedata.normalize("NFKD", text.casefold())
#
# def caseless_equal(left, right):
#     return normalize_caseless(left) == normalize_caseless(right)
#
# print(caseless_equal("james" ,"Jamess"))



# import re
# from openpyxl import load_workbook
# from collections import Counter
#
#
# def excelinput(filetoeopn, filecheckksheets, columnNo):
#     columnlist = []
#     # loading in workbook

#     wb = load_workbook(filetoeopn)
#
#     print("Sheets names:")
#     # obtaining sheets names
#     print(wb.get_sheet_names())
#     sheet = wb[wb.get_sheet_names()[filecheckksheets]]
#
#     #
#     data = sheet.values
#
#     for r in data:
#         # print(r[0])
#         if r[columnNo] is not None:
#             # print(r[columnNo])
#             columnlist.append(r[columnNo])
#
#     return columnlist
#
# for i in excelinput("datafiles\sgforums.xlsx", 0, 0):
#     if re.search("don",str(i))!="None":
#         print(i)
#
# x = '197a0"'
# x1 = "2)dsad"
# x2 = "3)dasasad"
# y = "3"
#
# print(x)
# print(re.match("^([\w])+\"$", x))

# x="asdsadsad fdsghgd haha"
# print(x.find("hahads"))

# if x.startswith('‘'):
#     print("work")




# y="text,,,0"
# x="(pp)"
# x1=""
# x2=""
# print(y.split(","))
#
# print(x[x.index(y.split(",")[1])+int(y.split(",")[3]):x.index(y.split(",")[2])])

# word = "James John"
# allword = word.split(" ")
# firstword=allword[0]
#
# x = ['James', "John", 'Peter', 'Le', 'has']
#
# # print(x.index("Peter"))
# if str(x[x.index(firstword)] + " " + x[x.index(firstword) + 1]) == word:
#     print(x[x.index(firstword)]+" "+x[x.index(firstword) + 1])
#     print(x.index(firstword))
#     print("Found")

# testarr = []
# stopword = ["a","i'll","the"]
#
# x = "whiat the games ass | a d asd the sd s/ads i'll fdsf ds.fcdsfs,     fsdf ds     fdsf"
#
# sentence=""
# for te in x.split(" "):
#     if str(te)!="":
#         if str(te)!=" ":
#             wordcount = 0
#             for qwe in stopword:
#
#                 print(te.strip())
#                 if str(re.match("^("+str(qwe)+")$", str(te).strip()))!="None":
#                     temp=str(te).replace(qwe, "")
#                     sentence=str(sentence)+" "+str(temp).strip()
#                     wordcount=wordcount+1
#             if wordcount==0:
#                 sentence = str(sentence) + " " + str(te).strip()
#             else:
#                 print("gdfg")
#
#
#
# print(sentence)



# for x1 in x.split(""):
#     if str(x1) != "":
#         if str(x1) != " ":
#             print(x1)
#             x1 = re.sub('\s+', ' ', x1).strip()
#             testarr.append(x1.strip())
#
# print(testarr)




# print(testarr)


# def excelinput(filetoeopn, filecheckksheets, columnNo):
#     columnlist = []
#     # loading in workbook
#     wb = load_workbook(filetoeopn)
#
#     print("Sheets names:")
#     # obtaining sheets names
#     print(wb.get_sheet_names())
#     print("Sheets Loaded: " + wb.get_sheet_names()[filecheckksheets])
#     sheet = wb[wb.get_sheet_names()[filecheckksheets]]
#
#     #
#     data = sheet.values
#
#     # print(next(data)[0:])
#
#     for r in data:
#         # print(r[0])
#         columnlist.append(r[columnNo])
#
#     return columnlist
#
# counttttt = 0
#
# currentFileNo=0
#
# for i in excelinput("datafiles\sgforums.xlsx", 0, 0):
#     if str(i) != "None":
#
#         print("Currently scanning Line: " + str(counttttt))
#         counttttt = counttttt + 1
#         i = str(i).replace('\n', " ")
#         i = str(i).replace('\\n', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#         i = str(i).replace('', " ")
#
#         i = str(i).replace('…', '...')
#         print(i)
#
#         f = open("..\sentencesInFiles\sentence_"+str(currentFileNo)+".txt","w", encoding="utf-8")
#         f.writelines(str(i))
#         f.close()
#         currentFileNo=int(currentFileNo)+1
#
# x = "0.78"
#
# x=float(x)+1
#
# print(x)

# x = "0/2"
#
# print(x[:str(x).index("/")])
# print(x[str(x).index("/")+1:])

# x="jamie le"
#
# print(x.fe le"
#
# print(x.find("jami4"))

# x="between marina bay"
# y="marina bay"
# print(y in x)

# print(os.getcwd())
# directory=os.getcwd()+"\Foo"
#
# if not os.path.exists(directory):
# try:
#
#     os.mkdir("Foo")
# except:
#     pass


# iam = " james"
# print(iam.split(""))

# import nltk
# nltk.download('stopwords')

# from nltk.corpus import stopwords
# print(stopwords.words('english'))

#
# import nltk
# nltk.download('reuters')


# print(str(re.search(r"\bam\b", str("am d"))))\pr
# nerln="jam i"
# jam="jam i ds"
#
# print(str(re.search("\\b(" + str(nerln) + ")\\b", str(jam).strip())) )
# import xlsxwriter as xlsxwriter
# from nltk import wordpunct_tokenize
# from nltk.corpus import stopwords
# from openpyxl import load_workbook
# import os
#
# word = ""
# inputExcelSheetForumData="..\datafiles\Edmwcompiled311017.xlsx"
#
# f = open('SinglishSentencesForNer.txt', 'w', encoding="utf-8")
#
#
# def excelinput(filetoeopn, filecheckksheets, columnNo):
#     columnlist = []
#     # loading in workbook
#     wb = load_workbook(filetoeopn)
#
#     print("Sheets names:")
#     # obtaining sheets names
#     print(wb.get_sheet_names())
#     sheet = wb[wb.get_sheet_names()[filecheckksheets]]
#
#     #
#     data = sheet.values
#
#     for r in data:
#         # print(r[0])
#         if r[columnNo] is not None:
#             # print(r[columnNo])
#             columnlist.append(str(r[columnNo]))
#
#     return columnlist
#
#
# def pythonFile(filetoopen, startreadArea, endReadArea, indextoadd, wordstoignore):
#     pyarr = []
#     f = open(filetoopen, 'r', encoding="utf8")
#     message = f.readlines()
#
#     # print(message)
#     f.close()
#     # print(message)
#     for myString in message:
#         # print(myString)
#
#         try:
#             if myString.find(wordstoignore) == -1:
#                 if startreadArea != "":
#                     pyarr.append(
#                         (str(myString[
#                              myString.index(startreadArea) + indextoadd:myString.index(endReadArea)]).lower().replace(
#                             "_", " ")))
#
#                 else:
#                     pyarr.append(myString.replace("\n", ""))
#         except:
#             pass
#     return pyarr
#
#     # read in
#
#     print(inread)
#
#
# senticNetWords = pythonFile("filesdb\senticnet.py", "['", "']", 2, "-----nil--------")
# # print(senticNetWords)
# print(set(senticNetWords))
#
# languages_ratios = {}
# row = 0
# col = 0
# workbook = xlsxwriter.Workbook('UserDatabase.xlsx')
# worksheet = workbook.add_worksheet()
#
# # worksheet.write(row, col, "ds")
# # row += 1
#
# # for inread in excelinput(inputExcelSheetForumData, 0, 0):
# tokens = wordpunct_tokenize(" dont bo chap leh ")
#
# words = [word.lower() for word in tokens]
#
# # list of sets of stopwords for each language
# dictOfSetsOfStopwords = {}
# languages_ratios = {}
#
# dictOfSetsOfStopwords["singlish"] = set(senticNetWords)
# for language in stopwords.fileids():
#     stopwords_set = set(stopwords.words(language))
#     dictOfSetsOfStopwords[language] = stopwords_set
#
# for eachLan in dictOfSetsOfStopwords:
#     # print(dictOfSetsOfStopwords[eachLan])
#     corpusSet = set(words)
#     common_elements = corpusSet.intersection(dictOfSetsOfStopwords[eachLan])
#     languages_ratios[eachLan] = len(common_elements)
#
# print(languages_ratios)
# print(languages_ratios["singlish"])
#
# if int(languages_ratios["singlish"]) > 2:
#     print("".join(inread.splitlines()))
#
#     f.writelines("".join(inread.splitlines()) + "\n\n")
#     worksheet.write(row, col, "".join(inread.splitlines()))
#     row += 1
# f.close()
# workbook.close()

# word = "going to"
# sentence="I am going to lan"
#
# print(str(re.search("\\b(" + str(word) + ")\\b", str(sentence))))
# from typing import Counter
# r = ["james","james"]
# print(Counter(r))



# from langdetect import detect
# from langdetect import detect_langs
# from langdetect import DetectorFactory
# DetectorFactory.seed = 0
#
# print(detect("James is good"))
# print(detect_langs("James is good"))

from langdetect import detect, DetectorFactory
from langdetect import detect_langs
DetectorFactory.seed = 0
# print(detect("My gpa cannot make it"))
print(detect_langs("419177 6117f"))

# print(detect_langs('What did you eat?'))

# from langdetect import detector_factory
# detector_factory.init_factory()
# print(detector_factory._factory.langlist)
# word ="  wrote"
#
#
# print(str(re.match("^([\w]+)\s+(wrote)$", str(word))))